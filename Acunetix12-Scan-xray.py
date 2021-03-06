import requests
import json
import tableprint
import os
import sys
import openpyxl as ws
from config import *

requests.packages.urllib3.disable_warnings()


def get_scans():
	creat_xlsx()
	c = 0
	while True:
		result = requests.get(define.host + "api/v1/scans?c=" + str(c), headers=define.api_header, timeout=30,
							  verify=False)
		results = json.loads(result.content)
		c = c + 100
		if results['scans'] == []:
			return print(define.RED + "[*]任务执行完毕 文件保存为:%s" % define.filename)
		for s in results["scans"]:
			get_vulnerabilities(s['scan_id'], s['current_session']['scan_session_id'], s['target']['address'])


def get_vulnerabilities(scan_id, scan_session_id, host):
	c = 0
	while True:
		result = requests.get(
			url=define.host + "api/v1/scans/" + scan_id + "/results/" + scan_session_id + "/vulnerabilities?c=%s" % str(
				c), headers=define.api_header, timeout=30, verify=False)
		results = json.loads(result.text)
		c = c + 100
		if results['vulnerabilities'] == []:
			return print(define.BLUE + "[-]当前扫描抓取结束 scan_id:%s" % scan_id)
		for s in results['vulnerabilities']:
			get_details(scan_id, scan_session_id, host, s['vuln_id'])


def get_details(scan_id, scan_session_id, host, vuln_id):
	vulnerabilities = {}
	result = requests.get(
		url=define.host + "api/v1/scans/" + scan_id + "/results/" + scan_session_id + "/vulnerabilities/" + vuln_id,
		headers=define.api_header, timeout=30, verify=False)
	results = json.loads(result.text)
	vulnerabilities['details'] = results['details']
	vulnerabilities['request'] = results['request']
	vulnerabilities['affects_url'] = results['affects_url']
	vulnerabilities['description'] = results['description']
	vulnerabilities['vt_name'] = results['vt_name']
	vulnerabilities['recommendation'] = results['recommendation']
	vulnerabilities['severity'] = results['severity']
	vulnerabilities['host'] = host
	vulnerabilities['affects_detail'] = results['affects_detail']
	write_xlsx(vulnerabilities)


def write_xlsx(vulnerabilities):
	print(define.GREEN + "[*]内容正在写入 vu_name:%s" % vulnerabilities['vt_name'])
	wb = ws.load_workbook(define.filename)
	sheet1 = wb['Sheet']
	num = sheet1.max_row
	sheet1.cell(row=num + 1, column=1, value=vulnerabilities['host'])
	sheet1.cell(row=num + 1, column=2, value=vulnerabilities['vt_name'])
	sheet1.cell(row=num + 1, column=3, value=vulnerabilities['severity'])
	sheet1.cell(row=num + 1, column=4, value=vulnerabilities['affects_detail'])
	sheet1.cell(row=num + 1, column=5, value=vulnerabilities['affects_url'])
	sheet1.cell(row=num + 1, column=6, value=vulnerabilities['request'])
	sheet1.cell(row=num + 1, column=7, value=vulnerabilities['recommendation'])
	sheet1.cell(row=num + 1, column=8, value=vulnerabilities['description'])
	sheet1.cell(row=num + 1, column=9, value=vulnerabilities['details'])
	wb.save(define.filename)


def creat_xlsx():
	if os.path.exists(define.filename) == False:
		s = 0
		wb = ws.Workbook()
		ws1 = wb.active
		if os.path.exists('out/') == False:
			os.mkdir('out')
		word = ['风险目标', '风险名称', '风险等级(3-0由高危到infomation)', '风险参数', '风险地址', '风险请求', '整改意见', '风险描述', '风险详情']
		for i in word:
			s = s + 1
			ws1.cell(row=1, column=s, value=i)
		wb.save(define.filename)
		print(define.RED + "[*]创建文件成功 %s" % define.filename)
	else:
		print(define.RED + "[*]文件已存在 文件为:%s" % define.filename)


x = []


def task(files):
	s = open('%s' % files, 'r')
	for i in s.readlines():
		i = i.strip()
		x.append(i)
	s.close

#设置代理的模块
def set_proxy(target_url,locationone):
    url = define.host + '/api/v1/targets/'+locationone+'/configuration'

    datajson = {
        "enabled": "true",
        "address": define.xray_address,
            "protocol": "http",
            "port": define.xray_port
    }

    datajsontwo = {
        "proxy": datajson
    }
    try:
        res = requests.patch(url, headers=define.api_header, verify=False,data=json.dumps(datajsontwo))
        if(res.status_code == 204):
            print("[+] " + target_url + " 代理设置成功")
    except:
        print("[-] " + target_url + " 代理设置失败")
        pass

def add_crawlonly(url):#仅爬取链接时，使用此模块对接xray
	# 添加任务
	data = {"address": url, "description": url, "criticality": "10"}
	try:
		response = requests.post(define.host + "api/v1/targets", data=json.dumps(data), headers=define.api_header,
								 timeout=30, verify=False)
		result = json.loads(response.content)
		try:
			respa = response.headers['Location']
		except:
			pass
		if "/api/v1/targets/" in respa:
			respa = respa.replace('/api/v1/targets/', '')
			set_proxy(url, respa)
		else:
			pass
		return result['target_id']
	except Exception as e:
		print(str(e))
		return

def add(url):#正常扫描时批量添加任务使用此模块
    #添加任务
    data = {"address":url,"description":url,"criticality":"10"}
    try:
        response = requests.post(define.host+"api/v1/targets",data=json.dumps(data),headers=define.api_header,timeout=30,verify=False)
        result = json.loads(response.content)
        return result['target_id']
    except Exception as e:
        print(str(e))
        return


def single_scan(url,scan):
	if scan == define.awvs_scan_rule['crawlonly']:
		target_id = add_crawlonly(url)
	else:
		target_id = add(url)
	data = {'target_id': target_id, 'profile_id': scan,
			'schedule': {'disable': False, 'start_date': None, 'time_sensitive': False}}
	try:
		r = requests.post(url=define.host + 'api/v1/scans', timeout=10, verify=False, headers=define.api_header,
						  data=json.dumps(data))
		if r.status_code == 201:
			print(define.BLUE + '[-] OK, 扫描任务已经启动 当前扫描:%s...' % url)
	except Exception as e:
		print(e)


def delete_all():
	c = 0
	print(define.RED + "[*]开始清除任务")
	while True:
		result = requests.get(define.host + "api/v1/targets?c=" + str(c), headers=define.api_header, timeout=30,
							  verify=False)
		results = json.loads(result.content)
		c = c + 100
		if results['targets'] == []:
			return print(define.RED + "[*]任务全部清除完毕")
		for s in results["targets"]:
			r = requests.delete(url=define.host + 'api/v1/targets/' + s['target_id'], timeout=10, verify=False,
								headers=define.api_header)
			print(define.BLUE + "[-]当前删除 target_id:%s" % s['target_id'])


if __name__ == '__main__':
	print(define.ORANGE + define.banner)
	if len(sys.argv) < 2:
		print(define.ORANGE + define.usage)
	elif sys.argv[1] == '-f':
		scan_list = [define.awvs_scan_rule['full'],define.awvs_scan_rule['highrisk'],define.awvs_scan_rule['XSS'],define.awvs_scan_rule['SQL'],
					 define.awvs_scan_rule['Weakpass'],define.awvs_scan_rule['crawlonly']]
		scan = input('请选择运行模式，1：full，2：highrisk，3：XSS，4：SQL，5：Weakpass，6：crawlonly\n')
		if scan.isdigit() == True and int(scan) <= 6 and  int(scan) != 0:
			scan = scan_list[int(scan) - 1]
			print(scan)
			try:
				task(str(sys.argv[2]))
				print(define.RED + "[*]扫描开始添加")
				for s in x:
					print(s)
					single_scan(s,scan)
				print(define.RED + "[*]扫描添加完毕")
			except:
				print(define.BLUE + '    [*]Usage example: Python3 Acunetix12-Scan-Agent.py -f url.txt')
		else:
			print('只能输入1到6的整数，请重新输入')
	elif sys.argv[1] == '-d':
		delete_all()
	elif sys.argv[1] == '-o':
		get_scans()
	else:
		print(define.ORANGE + define.usage)
